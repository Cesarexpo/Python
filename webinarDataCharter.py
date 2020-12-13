import openpyxl
import os
from datetime import date

os.chdir(r"C:\Users\coron\Desktop")
wb = openpyxl.load_workbook("Project Charter - Template.xlsx")
print(wb.sheetnames)
sheet = wb["Project Charter"]

# To create new spreadsheet
#wb.create_sheet(title="This is new")

# Webinar date value
today = date.today()
sheet["C3"].value = today
print(sheet["C3"].value)

# Webinar title value
sheet["C4"].value = "Charge Variant Webinar"
print(sheet["C4"].value)

# Job owner value
sheet["C5"].value = "Cesar Corona"
print(sheet["C5"].value)

# Engagement level value
sheet["C8"].value = "Toasty Warm"
print(sheet["C8"].value)

# Campaign value
sheet["C12"].value = "BIOLOGICS"
print(sheet["C12"].value)

# Engagement level value
sheet["C8"].value = "Toasty Warm"
print(sheet["C8"].value)

# Comment with webinar title, and column used
sheet["B21"].value = "Attended " + "IndustryTechnique " + "Webinar Title " + "On Date " + \
    str(today) + \
    " and indicates that their most used column is [insert from column P, if it's not other]"
print(sheet["B21"].value)

# Free Item Code value
sheet["B17"].value = "Free Item Code for this webinar"
print(sheet["B17"].value)

# Functional Role value
sheet["B27"].value = "As indicated in column [value]"
print(sheet["B27"].value)


wb.save("Project Charter - Toasty Warm.xlsx")
