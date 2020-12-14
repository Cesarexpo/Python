import openpyxl
import os
from datetime import date

os.chdir(r"C:\Users\coron\OneDrive\Atom\Python")
wb = openpyxl.load_workbook("Project Charter - Template.xlsx")
print(wb.sheetnames)
sheet = wb["Project Charter"]

webinar_title = "Troubleshooting for Gas Chromatography"
job_owner = "César Corona"
campaign_value = "LEADNURTURING"
date = "December 1, 2020"
free_item_code = "WBRGCTRBS20"
today = "December 7, 2020"


# To create new spreadsheet
#wb.create_sheet(title="This is new")

# Blazing Hot

# Submission value
# today = date.today()
sheet["C3"].value = today
print(sheet["C3"].value)

# Webinar title value
sheet["C4"].value = webinar_title
print(sheet["C4"].value)

# Job owner value
sheet["C5"].value = job_owner
print(sheet["C5"].value)

# Engagement level value
sheet["C8"].value = "Blazing Hot"
print(sheet["C8"].value)

# Campaign value
sheet["C12"].value = campaign_value
print(sheet["C12"].value)

# Comment with webinar title, and column used
sheet["B21"].value = "Attended webinar " + webinar_title + ", on " + date + \
    ". Requested quote. Is interested in products [as indicated in column N], and indicated that their most used column is [insert from column L, if it is not Other]"
print(sheet["B21"].value)

# Free Item Code value
sheet["B17"].value = free_item_code + "-Attended"
print(sheet["B17"].value)

# Functional Role value
sheet["B27"].value = "Job Role, as indicated in column [K]"
print(sheet["B27"].value)

# LC/MS question
sheet["B28"].value = "LC/MS question, as indicated in column [P]"
print(sheet["B28"].value)

wb.save("Project Charter - Blazing Hot.xlsx")


# Toasty Warm

# Submission value
# today = date.today()
sheet["C3"].value = today
print(sheet["C3"].value)

# Webinar title value
sheet["C4"].value = webinar_title
print(sheet["C4"].value)

# Job owner value
sheet["C5"].value = job_owner
print(sheet["C5"].value)

# Engagement level value
sheet["C8"].value = "Toasty Warm"
print(sheet["C8"].value)

# Campaign value
sheet["C12"].value = campaign_value
print(sheet["C12"].value)

# Comment with webinar title, and column used
sheet["B21"].value = "Attended webinar " + webinar_title + ", on " + date + \
    ". Is interested in products [as indicated in column N] and indicated that their most used column is [insert from column M, if it is not Other]"
print(sheet["B21"].value)

# Free Item Code value
sheet["B17"].value = free_item_code + "-Attended"
print(sheet["B17"].value)

# Functional Role value
sheet["B27"].value = "Job Role, as indicated in column [K]"
print(sheet["B27"].value)

# LC/MS question
sheet["B28"].value = "LC/MS question, as indicated in column [P]"
print(sheet["B28"].value)

wb.save("Project Charter - Toasty Warm.xlsx")


# New Project

# Submission value
# today = date.today()
sheet["C3"].value = today
print(sheet["C3"].value)

# Webinar title value
sheet["C4"].value = webinar_title
print(sheet["C4"].value)

# Job owner value
sheet["C5"].value = job_owner
print(sheet["C5"].value)

# Engagement level value
sheet["C8"].value = "New Project"
print(sheet["C8"].value)

# Campaign value
sheet["C12"].value = campaign_value
print(sheet["C12"].value)

# Comment with webinar title, and column used
sheet["B21"].value = "Registered for webinar " + webinar_title + ", broadcast on " + date + \
    ". Indicated that their most used column is [insert from column N, if it is not Other]"
print(sheet["B21"].value)

# Free Item Code value
sheet["B17"].value = "As indicated in column [A]"
print(sheet["B17"].value)

# Functional Role value
sheet["B27"].value = "Job Role, as indicated in column [L]"
print(sheet["B27"].value)

# LC/MS question
sheet["B28"].value = "LC/MS question, as indicated in column [Q]"
print(sheet["B28"].value)


wb.save("Project Charter - New Project.xlsx")
