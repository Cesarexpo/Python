# Operational version 20201213
import openpyxl
import os
from datetime import date
import pyodbc
import pandas as pd
[x for x in pyodbc.drivers() if x.startswith('Microsoft Access Driver')]


def createSpreadsheets(webinar_identifier):

    # MS Access Start
    # MS Access section: Retrieve variable values

    # define components of our connection string
    conn_str = (
        r'DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};'
        r'DBQ=C:\Users\coron\OneDrive\Atom\Python\Webinars.accdb;'
    )

    # create a connection to the database
    cnxn = pyodbc.connect(conn_str)

    # define the components of a query
    table_name = 'Automation'

    FreeItemCode = webinar_identifier

    # define query
    query = 'SELECT * FROM {}'.format(table_name)

    # define dataframe
    data = pd.read_sql(query, cnxn, index_col="FreeItemCode")

    # define selected FreeItemCode

    # Check data
    print(data.columns)
    print(data.head())

    # define variable values

    webinar_title = data.loc[webinar_identifier, "WebinarTitle"]
    brc = data.loc[webinar_identifier, "BRC"]
    # freeitemcode = data.loc[webinar_identifier, "FreeItemCode"]
    presenterfirstname = data.loc[webinar_identifier, "PresenterFirstName"]
    presenter = data.loc[webinar_identifier, "Presenter"]
    date = data.loc[webinar_identifier, "Date"]
    time = data.loc[webinar_identifier, "Time"]
    industry = data.loc[webinar_identifier, "Industry"]
    technique = data.loc[webinar_identifier, "Technique"]
    # registration_sales = data.loc[webinar_identifier, "Registration Sales"]
    overview = data.loc[webinar_identifier, "Overview"]
    registrantstotal = data.loc[webinar_identifier, "Registrants Total"]
    attendees = data.loc[webinar_identifier, "Attendees"]
    whoshouldattend = data.loc[webinar_identifier, "WhoShouldAttend"]
    learning_objectives = data.loc[webinar_identifier, "LearningObjectives"]
    # print(webinartitle, presenter, date, time, whoshouldattend, brc,webinar_identifier, overview, registrantstotal, attendees)
    # return webinartitle, presenter, date, time, whoshouldattend, brc, webinar_identifier, overview, registrantstotal, attendees

# MS Access End

    os.chdir(r"C:\Users\coron\OneDrive\Atom\Python")
    wb = openpyxl.load_workbook("Project Charter - Template.xlsx")
    print(wb.sheetnames)
    sheet = wb["Project Charter"]

    job_owner = "César Corona"
    campaign_value = "LEADNURTURING"
    free_item_code = webinar_identifier
    today = "December 7, 2020"

    # To create new spreadsheet
    #wb.create_sheet(title="This is new")

    # Blazing Hot

    # Submission value
    # today = date.today()
    sheet["C3"].value = today
    print(sheet["C3"].value)

    # Webinar title value
    sheet["C4"].value = webinar_title
    print(sheet["C4"].value)

    # Job owner value
    sheet["C5"].value = job_owner
    print(sheet["C5"].value)

    # Engagement level value
    sheet["C8"].value = "Blazing Hot"
    print(sheet["C8"].value)

    # Campaign value
    sheet["C12"].value = campaign_value
    print(sheet["C12"].value)

    # Comment with webinar title, and column used
    sheet["B21"].value = "[As indicated in column R]"
    print(sheet["B21"].value)

    # Free Item Code value
    sheet["B17"].value = free_item_code + "-Attended"
    print(sheet["B17"].value)

    # Functional Role value
    sheet["B27"].value = "Job Role, as indicated in column [K]"
    print(sheet["B27"].value)

    # LC/MS question
    sheet["B28"].value = "LC/MS question, as indicated in column [P]"
    print(sheet["B28"].value)

    wb.save("Project Charter - Blazing Hot.xlsx")

    # Toasty Warm

    # Submission value
    # today = date.today()
    sheet["C3"].value = today
    print(sheet["C3"].value)

    # Webinar title value
    sheet["C4"].value = webinar_title
    print(sheet["C4"].value)

    # Job owner value
    sheet["C5"].value = job_owner
    print(sheet["C5"].value)

    # Engagement level value
    sheet["C8"].value = "Toasty Warm"
    print(sheet["C8"].value)

    # Campaign value
    sheet["C12"].value = campaign_value
    print(sheet["C12"].value)

    # Comment with webinar title, and column used
    sheet["B21"].value = "[As indicated in column R]"
    print(sheet["B21"].value)

    # Free Item Code value
    sheet["B17"].value = free_item_code + "-Attended"
    print(sheet["B17"].value)

    # Functional Role value
    sheet["B27"].value = "Job Role, as indicated in column [K]"
    print(sheet["B27"].value)

    # LC/MS question
    sheet["B28"].value = "LC/MS question, as indicated in column [P]"
    print(sheet["B28"].value)

    wb.save("Project Charter - Toasty Warm.xlsx")

    # New Project

    # Submission value
    # today = date.today()
    sheet["C3"].value = today
    print(sheet["C3"].value)

    # Webinar title value
    sheet["C4"].value = webinar_title
    print(sheet["C4"].value)

    # Job owner value
    sheet["C5"].value = job_owner
    print(sheet["C5"].value)

    # Engagement level value
    sheet["C8"].value = "New Project"
    print(sheet["C8"].value)

    # Campaign value
    sheet["C12"].value = campaign_value
    print(sheet["C12"].value)

    # Comment with webinar title, and column used
    sheet["B21"].value = "[As indicated in column R]"
    print(sheet["B21"].value)

    # Free Item Code value
    sheet["B17"].value = "As indicated in column [A]"
    print(sheet["B17"].value)

    # Functional Role value
    sheet["B27"].value = "Job Role, as indicated in column [L]"
    print(sheet["B27"].value)

    # LC/MS question
    sheet["B28"].value = "LC/MS question, as indicated in column [Q]"
    print(sheet["B28"].value)

    wb.save("Project Charter - New Project.xlsx")
